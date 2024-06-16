import openpyxl
import openpyxl.chart
import openpyxl.chart.series

wb1 = openpyxl.load_workbook("octobre.xlsx",data_only=True)
wb2 = openpyxl.load_workbook("novembre.xlsx",data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx",data_only=True)

#{"Pommes": (760,660,900)}

def ajouter_data_depuis_wb(wb,d):
    sheet = wb.active
    for row in range(2,sheet.max_row): # type: ignore
        nom_article = sheet.cell(row,1).value # type: ignore
        if not nom_article:
            break
        total_ventes = sheet.cell(row,4).value # type: ignore
        if d.get(nom_article):
            d[nom_article].append(total_ventes)
        else:
            d[nom_article] = [total_ventes]

donnees = {}
ajouter_data_depuis_wb(wb1,donnees)
ajouter_data_depuis_wb(wb2,donnees)
ajouter_data_depuis_wb(wb3,donnees)


print(donnees)

wb_sortie = openpyxl.Workbook()
sheet = wb_sortie.active
sheet["A1"] = "Article" # type: ignore
sheet["B1"] = "Octobre" # type: ignore
sheet["C1"] = "Novembre" # type: ignore
sheet["D1"] = "Décembre" # type: ignore

row = 2
for i in donnees.items():
    nom_arcticle = i[0]
    ventes = i[1]
    sheet.cell(row,1).value = nom_arcticle # type: ignore
    for j in range(0,len(ventes)):
        sheet.cell(row, 2+j).value = ventes[j]# type: ignore
    row += 1
    
chart_ref = openpyxl.chart.Reference(sheet, min_col=2,min_row=2,max_col=sheet.max_column, max_row=2) # type: ignore
chart_serie = openpyxl.chart.Series(chart_ref,title="Total vente $") # type: ignore
chart = openpyxl.chart.BarChart()
chart.title = "Evolution du prix des pommes"
chart.append(chart_serie)

sheet.add_chart(chart) # type: ignore


wb_sortie.save("total_ventes_trimestre.xlsx")