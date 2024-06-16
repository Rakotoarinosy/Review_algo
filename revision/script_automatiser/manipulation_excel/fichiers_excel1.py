import openpyxl

wb = openpyxl.load_workbook("octobre.xlsx")
print(wb.sheetnames)
#sheet = wb[wb.sheetnames[0]]
sheet = wb.active

#sheet = wb["Feuil1"]
#cell = sheet["A1"] # type: ignore

print(sheet.max_row) # type: ignore
print(sheet.max_column) # type: ignore

# for row in range(2,7):
#     cell = sheet.cell(row,2) # type: ignore
#     print(cell.value)